#Marwenda version 1.00
#Facturas de Listado a modelo, NO se adjunta en el repositorio ningun tipo de informacion sensible
#Variables autoexplicativas(la mayoria)


#Creamos una execpcion por si falla la apertura de la libreria.
try:
    import sys
    from openpyxl.utils.cell import get_column_letter
    from openpyxl import Workbook, load_workbook
    from tkinter import *
except ImportError:
    # Que hacer si el módulo no se puede importar
    print("Módulo no instalado")

#Inicializamos el libro contable que nos interesa
wb = load_workbook('registro.xlsx')
#Inicializamos solo la hoja del libro que nos interesa
ws = wb['Facturas simpli']

#Creamos una funcion para que solo pasandole un numero me de los valores que esperamos de la venta
def Printarventa(NumeroFila):
   #Guardamos todos los valores que nos interesan en variables que usaremos despues.
    NumFact = str(ws['A' + NumeroFila].value)
    Fecha = str(ws['B' + NumeroFila].value)
    Euros = str(ws['H' + NumeroFila].value)
    Ip = str(ws['M' + NumeroFila].value)
    Pais = str(ws['N' + NumeroFila].value)
    IVA = str(ws['I' + NumeroFila].value)
    Nombre = str(ws['E' + NumeroFila].value)+" "+str(ws['F' + NumeroFila].value)
    Descripcion= str(ws['G' + NumeroFila].value)

#Abrimos el libro del modelo de factura, y escogemos la hoja de factura simplificada
    Wbmod = load_workbook('modelo.xlsx')
    WSmod = Wbmod['Factura Simpli']
#Rellenamos los valores que nos interesan del registro.
    WSmod['D5'].value = Fecha
    WSmod['D7'].value = NumFact
    WSmod['A11'].value = Nombre
    WSmod['A12'].value = Ip
    WSmod['A13'].value = Pais
    WSmod['A17'].value = Descripcion
    #Cambiamos el . por la , para que excell pueda trabajar con el numero
    WSmod['D17'].value = Euros.replace('.' , ',')

    #Comprobamos que el IVA sea distinto a 0, si no por defecto es 21%
    if(IVA == '0.0' ):
            WSmod['E23'].value = '0'
  
#Guardamos cada factura con su propio libro para subirlo a Drive o transformarlo en PDF a posteriori.

    Wbmod.save('Facturas/'+str(NumFact) + '.xlsx')

    # print( # ("Factura numero " + NumFact + " Fecha de Factura " + Fecha  
       # + " Cantidad pagada "  + Euros + " Euros " + "Nombre cliente "+ Nombre + " del pais "+ Pais +
        # " buscado por la Ip "+ Ip + " Se le grava un IVA de " + IVA + " Compro el video "+ Descripcion))





#Para recorrer todo el archivo sacando la informacion que necesitamos.
#Usamos el rango de 2,122 para sacar las primeras 121 Facturas simplificadas.
#Funciona perfectamente.

#for i in range(226,227): # ACUERDATE de poner un numero mas de la fila ultima del registro.
 #   Printarventa(str(i))


def sumar():
    try:
        _valor1 = int(entrada_texto.get())
        _valor2 = int(entrada2_texto.get())
        for i in range(_valor1,_valor2): # ACUERDATE de poner un numero mas de la fila ultima del registro.
           Printarventa(str(i))
        etiqueta4.config(text="Exito")
    except ValueError:
        etiqueta4.config(text="Introduce un numero")




app = Tk()
app.title("Facturas Simplif")

# Ventana Principal
vp = Frame(app)
vp.grid(column=0, row=0, padx=(50, 50), pady=(10, 10))
vp.columnconfigure(0, weight=1)
vp.rowconfigure(0, weight=1)

etiqueta = Label(vp, text="Desde la fila ")
etiqueta.grid(column=1, row=1, sticky=(W, E))
etiqueta2 = Label(vp, text="Hasta la fila")
etiqueta2.grid(column=3, row=1, sticky=(W, E))

etiqueta4= Label(vp, text="ACUERDATE de poner un numero mas de la fila ultima del registro.")
etiqueta4.grid(column=2, row=5, sticky=(W, E))

boton = Button(vp, text="Crear facturas", command=sumar)
boton.grid(column=1, row=4)

valor = "desde"
entrada_texto = Entry(vp, width=10, textvariable=valor)
entrada_texto.grid(column=2, row=1)
valor2 = "hasta"
entrada2_texto = Entry(vp, width=10, textvariable=valor2)
entrada2_texto.grid(column=4, row=1)


app.mainloop()
